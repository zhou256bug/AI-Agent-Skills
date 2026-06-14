#!/usr/bin/env python3
"""下载并解析第 23 期周报 Excel 附件 - 使用 openpyxl"""

import imaplib
import email
from email import policy
import ssl
import json
from openpyxl import load_workbook
from io import BytesIO

# 配置
IMAP_HOST = 'imap.qiye.aliyun.com'
IMAP_PORT = 993
IMAP_USER = 'qiang.zhou@newpostech.com'
IMAP_PASS = 'UIkm6I2cigBGA2Nl'
WEEK23_EMAIL_ID = '6219'

print("📋 开始解析第 23 期周报...\n")

ssl_context = ssl.create_default_context()
mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT, ssl_context=ssl_context)

try:
    mail.login(IMAP_USER, IMAP_PASS)
    mail.select('INBOX')
    
    status, msg_data = mail.fetch(WEEK23_EMAIL_ID, '(RFC822)')
    raw_email = msg_data[0][1]
    msg = email.message_from_bytes(raw_email, policy=policy.default)
    
    print(f"✅ 邮件：{msg.get('Subject', '')}")
    
    # 查找 Excel 附件
    print("\n🔍 查找 Excel 附件...")
    
    for part in msg.walk():
        content_disposition = str(part.get_content_disposition())
        
        if 'attachment' in content_disposition:
            filename = part.get_filename()
            if filename and ('23 期' in filename or '.xls' in filename.lower()):
                print(f"✅ 找到：{filename}")
                
                payload = part.get_payload(decode=True)
                if payload:
                    print(f"📊 大小：{len(payload)} bytes")
                    
                    try:
                        # 使用 openpyxl 读取 Excel
                        excel_file = BytesIO(payload)
                        wb = load_workbook(excel_file, data_only=True)
                        
                        print(f"\n📄 Excel 包含 {len(wb.sheetnames)} 个工作表:")
                        for i, sheet in enumerate(wb.sheetnames[:5], 1):
                            print(f"   {i}. {sheet}")
                        
                        # 读取第一个工作表
                        ws = wb.active
                        print(f"\n📊 工作表：{ws.title}")
                        print(f"   行数：{ws.max_row}")
                        print(f"   列数：{ws.max_column}")
                        
                        # 读取表头
                        headers = []
                        for col in range(1, min(ws.max_column + 1, 20)):
                            cell = ws.cell(row=1, column=col)
                            headers.append(str(cell.value) if cell.value else '')
                        
                        print(f"\n列名：{', '.join(headers[:10])}")
                        
                        # 识别人名和工作列
                        name_col_idx = None
                        work_col_idx = None
                        
                        name_keywords = ['姓名', '人员', '员工', 'Name', 'name']
                        work_keywords = ['工作内容', '工作', '本周工作', '总结', '计划', '进展']
                        
                        for i, h in enumerate(headers, 1):
                            if any(k in h for k in name_keywords):
                                name_col_idx = i
                            if any(k in h for k in work_keywords):
                                work_col_idx = i
                        
                        if name_col_idx and work_col_idx:
                            print(f"\n✅ 人员列：第{name_col_idx}列")
                            print(f"✅ 工作列：第{work_col_idx}列")
                            
                            # 提取人员周报
                            employees = []
                            for row in range(2, min(ws.max_row + 1, 50)):  # 最多 50 人
                                name_cell = ws.cell(row=row, column=name_col_idx)
                                work_cell = ws.cell(row=row, column=work_col_idx)
                                
                                name = str(name_cell.value) if name_cell.value else ''
                                work = str(work_cell.value) if work_cell.value else ''
                                
                                if name and len(name) > 1 and name not in ['nan', 'None', 'None']:
                                    employees.append({
                                        'name': name.strip(),
                                        'work': work.strip()[:500]
                                    })
                            
                            print(f"\n👤 共提取 {len(employees)} 人的周报")
                            
                            # 保存结果
                            result = {
                                'week_num': '23',
                                'email_subject': msg.get('Subject', ''),
                                'date': msg.get('Date', ''),
                                'employee_count': len(employees),
                                'employees': employees
                            }
                            
                            with open('/tmp/week23_parsed.json', 'w', encoding='utf-8') as f:
                                json.dump(result, f, ensure_ascii=False, indent=2)
                            
                            print("✅ 解析结果已保存到 /tmp/week23_parsed.json")
                            
                            # 生成移动端汇报
                            print("\n" + "=" * 40)
                            print("📋 工作周报 · 第 23 期")
                            print("📧 发件人：林珊雯")
                            print(f"👤 共{len(employees)}人")
                            print("━" * 40)
                            
                            for emp in employees[:10]:
                                print(f"\n👤 {emp['name']}")
                                print("─" * 15)
                                work_preview = emp['work'][:100]
                                if len(emp['work']) > 100:
                                    work_preview += '…'
                                print(work_preview)
                            
                            if len(employees) > 10:
                                print(f"\n… 还有 {len(employees) - 10} 人")
                            
                            print("\n" + "━" * 40)
                            print("✅ 周报整理完成")
                            
                        else:
                            print("⚠️  未能自动识别人员列和工作列")
                            print(f"   可用列：{headers}")
                        
                    except Exception as e:
                        print(f"❌ 解析 Excel 失败：{e}")
                        import traceback
                        traceback.print_exc()
                break
    
    mail.logout()
    
except Exception as e:
    print(f"❌ 错误：{e}")
    import traceback
    traceback.print_exc()
    try:
        mail.logout()
    except:
        pass
